#! /usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("merged.xlsx")
sheet = wb["Sheet"]
sheet.unmerge_cells("A1:D3")
sheet.unmerge_cells("C5:D5")
wb.save("unmerged.xlsx")
