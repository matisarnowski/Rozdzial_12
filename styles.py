#! /bin/python3

import openpyxl
from openpyxl.styles import Font, NamedStyle

wb = openpyxl.Workbook()
sheet = wb["Sheet"]

font_times_new_roman = Font(name="Times New Roman", bold=True)
styleObj1 = NamedStyle(name="style_Obj1", font=font_times_new_roman)
wb.add_named_style(styleObj1)
sheet.cell(1, 1).style = "style_Obj1"
sheet["A1"] = "Pogrubiona czcionka Times New Roman"

font_calibri = Font(size=24, italic=24)
styleObj2 = NamedStyle(name="styleObj2", font=font_calibri)
wb.add_named_style(styleObj2)
sheet.cell(3, 2).style = "styleObj2"
sheet["B3"] = "Pochylona czcionka o wielkości 24 punktów."

wb.save("stylesCorrected.xlsx")
