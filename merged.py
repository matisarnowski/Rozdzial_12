#! /usr/bin/python3

import openpyxl

wb = openpyxl.Workbook()
sheet = wb["Sheet"]
sheet.merge_cells("A1:D3")
sheet["A1"] = "Dwanaście komórek połączonych ze sobą."
sheet.merge_cells("C5:D5")
sheet["C5"] = "Dwie połączone komórki"
wb.save("mergedCorrected.xlsx")
